import io
from dataclasses import dataclass
from dataclasses import asdict
import datetime
from typing import Any, Dict, List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import polars as pl

from apps.settings.configs import check_lang_jn_in_df
from apps.settings.configs import JnDataCols
from apps.settings.configs import rename_en_to_jn_in_df
from apps.settings.configs import XlsSummaryConfs
from apps.settings.configs import XlsResultConfs
from apps.settings.configs import XlsDetailConfs


@dataclass
class Summary:
    office: str
    branch_office: str
    local_area: str
    address: str
    person: str
    reciver: str
    projective_name: int
    start: str
    end: str
    pt_count: int
    signal_frec: int
    area: float
    outline_length: float
    min_epochs: int
    max_pdop: float
    min_satellites: int
    work_days: str
    work_time: str
    dict = asdict

@dataclass
class WorkTime:
    days: str
    work_time: str


class WriteExcelFile(XlsResultConfs):
    def __init__(self, df: pl.DataFrame, area: float, length: float):
        super().__init__()
        self.jn_confs = JnDataCols()
        self.summary_confs = XlsSummaryConfs()
        self.df = self._str_to_datetime(self._rename(df))
        self.summary = self.get_summary(df, area, length)
        self.wb = openpyxl.open(self.summary_confs.template_file)
        self.main_ws = self.wb[self.summary_confs.sheet_name]

    def _rename(self, df):
        if check_lang_jn_in_df(df):
            return df
        else:
            rename_en_to_jn_in_df(df)
    
    def _str_to_datetime(self, df: pl.DataFrame) -> pl.DataFrame:
        fmt = '%Y-%m-%d %H:%M:%S'
        if df[self.jn_confs.start_datetime_col].dtype != pl.Datetime:
            df = df.with_columns([
                pl.col(self.jn_confs.start_datetime_col).str.strptime(pl.Datetime, fmt)
            ])
        if df[self.jn_confs.datetime_col].dtype != pl.Datetime:
            df = df.with_columns([
                pl.col(self.jn_confs.datetime_col).str.strptime(pl.Datetime, fmt)
            ])
        return df

    def _get_start_and_end(self) -> Dict[str, str]:
        fmt = '%Y-%m-%d %H:%M:%S'
        start = self.df[self.jn_confs.start_datetime_col].min().strftime(fmt)
        end = self.df[self.jn_confs.datetime_col].max().strftime(fmt)
        return {'start': start, 'end': end}

    def _timedelta_to_str(self, delta: datetime.timedelta) -> str:
        days = delta.days
        hours = delta.seconds // 3600
        if 0 < hours:
            minutes = int(round((delta.seconds - hours * 3600) / 60, 0))
        else:
            minutes = int(round(delta.seconds / 60, 0))
        if days == 0:
            return f"{hours}h {minutes}m"
        else:
            return f"{days}d {hours}h {minutes}m"

    @property
    def _work_time(self) -> WorkTime:
        starts = self.df[self.jn_confs.start_datetime_col].to_list()
        ends = self.df[self.jn_confs.datetime_col].to_list()
        ts = pl.DataFrame({'datetime': starts + ends})
        df = (
            ts.with_columns([
                pl.col('datetime').dt.date().alias('date')
            ])
            .group_by(['date'])
                .agg([
                    pl.col('datetime').min().alias('start'),
                    pl.col('datetime').max().alias('end'),
                ])
            .with_columns([
                (pl.col('end') - pl.col('start')).alias('work_time')
            ])
        )
        days = f"{len(df['date'])} 日"
        time = self._timedelta_to_str(df['work_time'].sum())
        return WorkTime(days, time)

    @property
    def __get_group_names(self) -> str:
        names = '（'
        for name in self.df['group'].unique().sort():
            if name == '':
                continue
            elif names[-1] != "（":
                names += ', '
            names += f"{name}"

        if names != '（':
            names += '）'
            return names
        else:
            return ''

    def get_summary(self, df: pl.DataFrame, area: float, length: float):
        confs = JnDataCols()
        ts = self._get_start_and_end()
        work_time = self._work_time
        summary = Summary(
            df[confs.office_col][0],
            df[confs.branch_office_col][0],
            df[confs.lcoal_area_col][0],
            df[confs.address_col][0] + self.__get_group_names,
            df[confs.person_col][0],
            df[confs.reciver_col][0],
            df[confs.epsg_col][0],
            ts.get('start'),
            ts.get('end'),
            df.shape[0],
            df.filter(2 <= pl.col(confs.signal_frec_col)).shape[0],
            area,
            length,
            df[confs.epochs_col].min(),
            df[confs.pdop_col].max(),
            df[confs.satellites_col].min(),
            work_time.days,
            work_time.work_time
        )
        return summary

    @property
    def __get_summary_coords_dict(self) -> Dict[str, Dict[str, int]]:
        coords_dict = (
            XlsSummaryConfs()
            .__dict__
            .get('_confs')
            .get('main_sheet')
            .get('summary_cells')
        )
        return coords_dict
    
    @property
    def write_summary_xl(self):
        # 概要をメインシートに入力する
        coords_dict = self.__get_summary_coords_dict
        summary_dict = self.summary.dict()
        for key, val in summary_dict.items():
            coords  = coords_dict.get(key)
            if coords is None:
                continue
            coords['value'] = val
            r = self.main_ws.cell(**coords)

    def _write_series_xl(
        self, 
        ws: Worksheet, 
        values: List[Any], 
        start_row: int, 
        start_col: int
    ):
        for value in values:
            _ = ws.cell(start_row, start_col, value)
            start_row += 1
    
    def _rewrite_label(self, label: str) -> str:
        find = '.'
        if find in label:
            if label[label.find(find): ] == '.0':
                return label[: label.find(find)]
        return label

    @property
    def write_point_num_xl(self):
        # Excelファイルに順番を書き込む
        coords = self.coords_pt_num_first
        length = len(self.df[self.jn_confs.lon_col].to_list())
        writes = [i + 1 for i in range(length)]
        self._write_series_xl(self.main_ws, writes, coords.row, coords.column)

    @property
    def write_point_names_xl(self):
        # Excelファイルに測点番号と測点名を書き込む
        coords = self.coords_pt_name_first
        names = self.df[self.jn_confs.pt_name_col].to_list()
        writes = []
        for name in names:
            name = self._rewrite_label(str(name)).replace(' ', '')
            writes.append(name)
        self._write_series_xl(self.main_ws, writes, coords.row, coords.column)

    @property
    def write_lons_xl(self):
        # Excelファイルに経度を書き込む
        coords = self.coords_lon_first
        writes = self.df[self.jn_confs.lon_col].to_list()
        self._write_series_xl(self.main_ws, writes, coords.row, coords.column)

    @property
    def write_lats_xl(self):
        # Excelファイルに緯度を書き込む
        coords = self.coords_lat_first
        writes = self.df[self.jn_confs.lat_col].to_list()
        self._write_series_xl(self.main_ws, writes, coords.row, coords.column)

    @property
    def write_epochs_xl(self):
        # Excelファイルに測定回数を書き込む
        coords = self.coords_epochs_first
        writes = self.df[self.jn_confs.epochs_col].to_list()
        self._write_series_xl(self.main_ws, writes, coords.row, coords.column)

    @property
    def write_pdop_xl(self):
        # ExcelファイルにPDOPを書き込む
        coords = self.coords_pdop_first
        writes = self.df[self.jn_confs.pdop_col].to_list()
        self._write_series_xl(self.main_ws, writes, coords.row, coords.column)

    @property
    def write_satellites_xl(self):
        # Excelファイルに衛星数を書き込む
        coords = self.coords_satellites_first
        writes = self.df[self.jn_confs.satellites_col].to_list()
        self._write_series_xl(self.main_ws, writes, coords.row, coords.column)

    @property
    def write_y_xl(self):
        # ExcelファイルにY（lon）を書き込む
        coords = self.coords_y_first
        col = self.jn_confs.y_col
        writes = None
        if col  in self.df.columns:
            writes = self.df[col].to_list()
        if (not writes is None) & (writes[0] != ''):
            writes = [round(c, 2) for c in writes]
            self._write_series_xl(self.main_ws, writes, coords.row, coords.column)
    
    @property
    def write_x_xl(self):
        # ExcelファイルにX（lat）を書き込む
        coords = self.coords_x_first
        col = self.jn_confs.x_col
        writes = None
        if col  in self.df.columns:
            writes = self.df[col].to_list()
        if (not writes is None) & (writes[0] != ''):
            writes = [round(c, 2) for c in writes]
            self._write_series_xl(self.main_ws, writes, coords.row, coords.column)
    
    @property
    def __side(self) -> openpyxl.styles.Side:
        # セルのスタイルを取得
        side = (
            openpyxl
            .styles
            .Side(style='thin', color='000000')
        )
        return side

    @property
    def __border(self) -> openpyxl.styles.Border:
        # 罫線のスタイルを取得
        side = self.__side
        border = (
            openpyxl
            .styles
            .Border(
                top=side, 
                bottom=side, 
                left=side, 
                right=side
            )
        )
        return border

    def write_border_line_xl(
        self, 
        ws: Worksheet, 
        start_row: int, 
        end_row: int, 
        end_col: int
    ):
        # 罫線を引く
        border = self.__border
        generator = (
            ws.iter_rows(
                min_row=start_row, 
                max_row=end_row, 
                max_col=end_col
            )
        )
        for row in generator:
            for cell in row:
                cell.border = border
    
    def __font(self, cell: openpyxl.cell.cell.Cell, font: str) -> openpyxl.styles.Font:
        font = (
            openpyxl
            .styles
            .Font(
                name='游ゴシック', 
                size=cell.font.size, 
                bold=cell.font.bold, 
                italic=cell.font.italic, 
                underline=cell.font.underline, 
                strike=cell.font.strike, 
                color=cell.font.color
            )
        )
        return font

    def set_font(
        self, 
        ws: Worksheet, 
        end_row: int, 
        end_col: int, 
        font: str='游ゴシック'
    ):
        for row in ws.iter_rows(min_row=1, max_row=end_row, max_col=end_col):
            for cell in row:
                cell.font = self.__font(cell, font)

    @property
    def write_main_page_xl(self):
        # メインページに測量成果を書き込む
        self.write_summary_xl
        self.write_point_num_xl
        self.write_point_names_xl
        self.write_lons_xl
        self.write_lats_xl
        self.write_epochs_xl
        self.write_pdop_xl
        self.write_satellites_xl
        self.write_y_xl
        self.write_x_xl
        # 罫線を引く
        start = self.coords_pt_num_first
        end = self.coords_x_first
        end.row += self.df.shape[0] - 1
        end.column += 1
        self.write_border_line_xl(self.main_ws, start.row, end.row, end.column)
        # Fontを強制的に"游ゴシック"に（テンプレートを調整してもなぜか"UDEV Gothic"になるので）
        self.set_font(self.main_ws, end.row, end.column)

    def write_details_columns_xl(self, ws: Worksheet, start_row: int, start_col: int):
        for i, val in enumerate(self.df.columns):
            col = i + start_col
            cell = ws.cell(row=start_row, column=col, value=val)

    def write_details_data_xl(self, ws: Worksheet, start_row, start_col: int):
        for i, data in enumerate(self.df.iter_rows()):
            row = i + start_row
            for j, val in enumerate(data):
                col = j + start_col
                cell = ws.cell(row=row, column=col, value=val)

    @property
    def write_details_page_xl(self):
        # 詳細ページにデータを書き込む
        conf = XlsDetailConfs()
        ws = self.wb[conf.sheet_name]
        start = conf.coords_cells_start
        self.write_details_columns_xl(ws, start.row, start.column)
        start.row += 1
        self.write_details_data_xl(ws, start.row, start.column)
        end_row, end_column = self.df.shape
        end_row = end_row + start.row - 1
        self.write_border_line_xl(ws, start.row - 1, end_row, end_column)
        self.set_font(ws, end_row, end_column)



def write_dataframe_to_xls_bytes(df: pl.DataFrame, area, length):
    # テンプレートにDataFrameの値をBufferに書き込み取得
    wex = WriteExcelFile(df, area, length)
    wex.write_main_page_xl
    wex.write_details_page_xl
    buffer = io.BytesIO()
    wex.wb.save(buffer)
    xls_bytes = buffer.getvalue()
    wex.wb.close()
    return xls_bytes